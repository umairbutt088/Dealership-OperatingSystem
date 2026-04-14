"""Per-vehicle workbook tabs: import/export `VehicleLineItem` rows."""

from __future__ import annotations

import re
from typing import Any, Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from dealershipos.db.models import Vehicle, VehicleLineItem
from dealershipos.services.plates import norm_plate

CORE_SHEETS: frozenset[str] = frozenset(
    {
        "Front Sheet",
        "Sold Stock",
        "Stock Data",
        "Collection",
        "Investor Budget",
        "SOR",
        "Investor Car Expense",
        "Expense",
        "Fuel Expense",
        "Money in",
        "Cash Spending",
        "Money Out",
    }
)


def _cell_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _header_map(ws, max_scan: int = 30) -> tuple[dict[str, int], int]:
    best_row = None
    best_count = 0
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        names = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for n in names if n)
        if non_empty >= 2:
            key_set = {_norm_header(n) for n in names if n}
            markers = {"item", "amount", "platenumber", "month", "investors"}
            if key_set & markers:
                if non_empty > best_count:
                    best_count = non_empty
                    best_row = (ri, names)
    if not best_row:
        return {}, -1
    ri, names = best_row
    hmap: dict[str, int] = {}
    for i, n in enumerate(names):
        if n:
            hmap[_norm_header(n)] = i
    return hmap, ri


def _vehicle_for_sheet_name(db: Session, sheet_name: str) -> Vehicle | None:
    ns = "".join(c for c in sheet_name.upper() if c.isalnum())
    if len(ns) < 5:
        return None
    for v in db.scalars(select(Vehicle)):
        vpn = v.plate_norm or norm_plate(v.plate)
        if len(vpn) >= 5 and vpn in ns:
            return v
    return None


def import_vehicle_line_sheets(wb: Workbook, db: Session) -> int:
    total = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in CORE_SHEETS:
            continue
        v = _vehicle_for_sheet_name(db, sheet_name)
        if not v:
            continue
        ws = wb[sheet_name]
        db.execute(
            delete(VehicleLineItem).where(
                VehicleLineItem.stock_id == v.stock_id,
                VehicleLineItem.sheet_name == sheet_name,
            )
        )
        hmap, hri = _header_map(ws)
        start_row = hri + 2 if hri >= 0 else 2
        if "item" in hmap and "amount" in hmap:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                i_idx, a_idx = hmap["item"], hmap["amount"]
                lab = row[i_idx] if i_idx < len(row) else None
                amt = row[a_idx] if a_idx < len(row) else None
                if lab is None and amt is None:
                    continue
                db.add(
                    VehicleLineItem(
                        stock_id=v.stock_id,
                        sheet_name=sheet_name,
                        item_label=str(lab) if lab is not None else "",
                        amount=_cell_float(amt),
                    )
                )
                total += 1
        else:
            for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
                if not row or len(row) < 2:
                    continue
                a, b = row[0], row[1]
                if a is None or str(a).strip() == "":
                    continue
                if isinstance(b, (int, float)):
                    pass
                elif isinstance(b, str) and b.strip():
                    bs = b.strip().replace(",", "")
                    if not bs.replace(".", "", 1).replace("-", "").isdigit():
                        continue
                else:
                    continue
                db.add(
                    VehicleLineItem(
                        stock_id=v.stock_id,
                        sheet_name=sheet_name,
                        item_label=str(a).strip(),
                        amount=_cell_float(b),
                    )
                )
                total += 1
    return total


def _safe_sheet_title(name: str) -> str:
    return (name[:31] if len(name) > 31 else name) or "Sheet"


def export_vehicle_line_sheets(wb: Workbook, db: Session) -> int:
    pairs = db.execute(
        select(VehicleLineItem.stock_id, VehicleLineItem.sheet_name).distinct()
    ).all()
    n_out = 0
    for stock_id, sheet_name in pairs:
        if not sheet_name:
            continue
        items = list(
            db.scalars(
                select(VehicleLineItem)
                .where(
                    VehicleLineItem.stock_id == stock_id,
                    VehicleLineItem.sheet_name == sheet_name,
                )
                .order_by(VehicleLineItem.id)
            )
        )
        if not items:
            continue
        title = _safe_sheet_title(sheet_name)
        if title in wb.sheetnames:
            ws = wb[title]
        else:
            ws = wb.create_sheet(title)
        for r in range(1, min(ws.max_row, 500) + 1):
            for c in range(1, 5):
                cell = ws.cell(r, c)
                if getattr(cell, "data_type", None) != "f":
                    cell.value = None
        ws.cell(1, 1, "Item")
        ws.cell(1, 2, "Amount")
        for i, li in enumerate(items, start=2):
            ws.cell(i, 1, li.item_label)
            ws.cell(i, 2, li.amount)
        n_out += 1
    return n_out
