"""Import/export aligned to Master_Spreadsheet workbook sheet names and headers."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils.datetime import from_excel
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from dealershipos.db.models import (
    CashSpendingRow,
    CollectionRow,
    DeliveryRow,
    ExpenseRow,
    FuelExpenseRow,
    Investor,
    InvestorCarExpenseRow,
    MonthlySummary,
    MoneyInRow,
    MoneyOutRow,
    SorRow,
    Vehicle,
    VehicleLineItem,
    new_stock_id,
)
from dealershipos.services.folders import ensure_car_folders
from dealershipos.services.plates import norm_plate


def _cell_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _cell_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _month_date(val: Any) -> Optional[date]:
    d = _cell_date(val)
    if d:
        return date(d.year, d.month, 1)
    return None


def _header_map(ws, max_scan: int = 30) -> tuple[dict[str, int], int]:
    """Return {normalized header: col index} and header row index."""
    best_row = None
    best_count = 0
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        names = []
        for c in row:
            if c is None:
                names.append("")
            else:
                names.append(str(c).strip())
        non_empty = sum(1 for n in names if n)
        if non_empty >= 3:
            key_set = {_norm_header(n) for n in names if n}
            markers = {
                "platenumber",
                "numberplatereference",
                "month",
                "make&model",
                "investors",
                "source",
                "category",
                "item",
                "datewon",
                "stockid",
            }
            if key_set & markers:
                if non_empty > best_count:
                    best_count = non_empty
                    best_row = (ri, names)
    if not best_row:
        return {}, -1
    ri, names = best_row
    hmap: dict[str, int] = {}
    for i, n in enumerate(names):
        if n:
            hmap[_norm_header(n)] = i
    return hmap, ri


def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _row_dict(row: tuple[Any, ...], hmap: dict[str, int], keys: list[str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k in keys:
        nk = _norm_header(k)
        if nk not in hmap:
            continue
        out[k] = row[hmap[nk]] if hmap[nk] < len(row) else None
    return out


def _get_or_create_vehicle(db: Session, plate_raw: str | None) -> Vehicle | None:
    np = norm_plate(plate_raw or "")
    if not np:
        return None
    r = db.execute(select(Vehicle).where(Vehicle.plate_norm == np)).scalar_one_or_none()
    if r:
        return r
    v = Vehicle(stock_id=new_stock_id(), plate=(plate_raw or "").strip().upper() or np, plate_norm=np)
    db.add(v)
    db.flush()
    ensure_car_folders(v.stock_id)
    return v


def _resolve_vehicle_from_row(
    db: Session,
    row: tuple[Any, ...],
    hmap: dict[str, int],
    plate_keys: tuple[str, ...] = ("platenumber", "numberplatereference"),
) -> Vehicle | None:
    """Prefer **Stock ID** column when present (`STK-...`), else plate columns."""
    if "stockid" in hmap:
        idx = hmap["stockid"]
        if idx < len(row):
            raw = row[idx]
            if raw:
                s = str(raw).strip().upper()
                if s.startswith("STK-"):
                    v = db.get(Vehicle, s)
                    if v:
                        return v
    for pk in plate_keys:
        if pk not in hmap:
            continue
        col = hmap[pk]
        if col >= len(row):
            continue
        plate = row[col]
        if plate is None or str(plate).strip() == "":
            continue
        return _get_or_create_vehicle(db, str(plate))
    return None


def import_workbook(path: Path | str, db: Session, replace: bool = False) -> dict[str, int]:
    """Import rows from a Master Spreadsheet–style workbook. Returns counts per section."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    counts: dict[str, int] = {}

    if replace:
        for model in (
            VehicleLineItem,
            InvestorCarExpenseRow,
            SorRow,
            FuelExpenseRow,
            CashSpendingRow,
            MoneyOutRow,
            MoneyInRow,
            ExpenseRow,
            CollectionRow,
            DeliveryRow,
            MonthlySummary,
            Vehicle,
            Investor,
        ):
            db.execute(delete(model))
        db.commit()

    # --- Stock Data ---
    if "Stock Data" in wb.sheetnames:
        ws = wb["Stock Data"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                v = _resolve_vehicle_from_row(db, row, hmap, ("platenumber",))
                if v is None:
                    continue
                if "month" in hmap:
                    v.month = _month_date(row[hmap["month"]]).isoformat() if _month_date(row[hmap["month"]]) else None
                if "dateaquired" in hmap or "dateacquired" in hmap:
                    ci = hmap.get("dateaquired") or hmap.get("dateacquired")
                    if ci is not None:
                        v.date_acquired = _cell_date(row[ci])
                if "make&model" in hmap:
                    m = row[hmap["make&model"]]
                    v.model = str(m) if m is not None else None
                inv_key = "investorsa" if "investorsa" in hmap else "investor/mp"
                if inv_key not in hmap:
                    inv_key = next((k for k in hmap if "investor" in k), None)
                if inv_key:
                    x = row[hmap[inv_key]]
                    v.investor = str(x) if x is not None else None
                if "source" in hmap:
                    x = row[hmap["source"]]
                    v.source = str(x) if x is not None else None
                if "pxvalue" in hmap:
                    v.px_value = _cell_float(row[hmap["pxvalue"]])
                if "price" in hmap:
                    v.purchase_price = _cell_float(row[hmap["price"]])
                rc_key = "reconditioningcosts" if "reconditioningcosts" in hmap else None
                if rc_key:
                    v.recon_cost = _cell_float(row[hmap[rc_key]])
                if "totalcost" in hmap:
                    v.total_cost = _cell_float(row[hmap["totalcost"]])
                st_key = "status" if "status" in hmap else None
                status = None
                if st_key:
                    s = row[hmap["status"]]
                    status = str(s) if s is not None else None
                    v.status = status
                sold_val = row[hmap["sold"]] if "sold" in hmap else None
                prof_val = row[hmap["profit"]] if "profit" in hmap else None
                if sold_val not in (None, "", "-"):
                    try:
                        v.sold_price = float(sold_val)
                    except (TypeError, ValueError):
                        pass
                if prof_val not in (None, "", "-"):
                    try:
                        v.profit = float(prof_val)
                    except (TypeError, ValueError):
                        pass
                if status and "sold" in status.lower():
                    v.is_sold = True
                elif v.sold_price and v.sold_price > 0:
                    v.is_sold = True
                else:
                    v.is_sold = False
                n += 1
            counts["stock_data"] = n

    # --- Sold Stock ---
    if "Sold Stock" in wb.sheetnames:
        ws = wb["Sold Stock"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                v = _resolve_vehicle_from_row(db, row, hmap, ("numberplatereference", "platenumber"))
                if v is None:
                    continue
                v.is_sold = True
                if "month" in hmap:
                    md = _month_date(row[hmap["month"]])
                    if md:
                        v.month = md.isoformat()
                if "dateaquired" in hmap or "dateacquired" in hmap:
                    ci = hmap.get("dateaquired") or hmap.get("dateacquired")
                    if ci is not None:
                        v.date_acquired = _cell_date(row[ci])
                if "make&model" in hmap:
                    m = row[hmap["make&model"]]
                    v.model = str(m) if m is not None else v.model
                inv_k = "sainvestorname" if "sainvestorname" in hmap else None
                if not inv_k:
                    inv_k = next((k for k in hmap if "investor" in k), None)
                if inv_k:
                    x = row[hmap[inv_k]]
                    v.investor = str(x) if x is not None else v.investor
                if "totalcost" in hmap:
                    v.total_cost = _cell_float(row[hmap["totalcost"]])
                if "sold" in hmap:
                    v.sold_price = _cell_float(row[hmap["sold"]])
                if "partex" in hmap:
                    v.part_ex = _cell_float(row[hmap["partex"]])
                if "totalprofit" in hmap:
                    v.profit = _cell_float(row[hmap["totalprofit"]])
                if "investorprofit" in hmap:
                    v.investor_profit = _cell_float(row[hmap["investorprofit"]])
                if "saprofit" in hmap:
                    v.sa_profit = _cell_float(row[hmap["saprofit"]])
                if "datelisted" in hmap:
                    v.date_listed = _cell_date(row[hmap["datelisted"]])
                if "datesold" in hmap:
                    v.date_sold = _cell_date(row[hmap["datesold"]])
                if "platform" in hmap or "platfrom" in hmap:
                    pk2 = "platform" if "platform" in hmap else "platfrom"
                    x = row[hmap[pk2]]
                    v.platform = str(x) if x is not None else None
                if "invoicenumber" in hmap:
                    x = row[hmap["invoicenumber"]]
                    v.invoice_number = str(x) if x is not None else None
                if "customername" in hmap:
                    x = row[hmap["customername"]]
                    v.customer_name = str(x) if x is not None else None
                if "contactinfo" in hmap:
                    x = row[hmap["contactinfo"]]
                    v.contact_info = str(x) if x is not None else None
                if "warranty" in hmap:
                    x = row[hmap["warranty"]]
                    v.warranty = str(x) if x is not None else None
                if "autoguardnumber" in hmap:
                    x = row[hmap["autoguardnumber"]]
                    v.autoguard = str(x) if x is not None else None
                n += 1
            counts["sold_stock"] = n

    # --- Investor Budget ---
    if "Investor Budget" in wb.sheetnames:
        ws = wb["Investor Budget"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                name_i = hmap.get("investors")
                if name_i is None:
                    continue
                name = row[name_i]
                if name is None or str(name).strip() == "":
                    continue
                inv = db.execute(select(Investor).where(Investor.name == str(name).strip())).scalar_one_or_none()
                if not inv:
                    inv = Investor(name=str(name).strip())
                    db.add(inv)
                inv.initial_balance = _cell_float(row[hmap["initialbalance"]]) if "initialbalance" in hmap else inv.initial_balance
                inv.capital_returned = _cell_float(row[hmap["capitalreturned"]]) if "capitalreturned" in hmap else inv.capital_returned
                inv.total_balance = _cell_float(row[hmap["totalbalance"]]) if "totalbalance" in hmap else inv.total_balance
                inv.purchased = _cell_float(row[hmap["purchased"]]) if "purchased" in hmap else inv.purchased
                tp_key = "totalprofitsincenov25" if "totalprofitsincenov25" in hmap else None
                if not tp_key:
                    tp_key = next((k for k in hmap if "totalprofit" in k), None)
                if tp_key:
                    inv.total_profit = _cell_float(row[hmap[tp_key]])
                if "available" in hmap:
                    inv.available = _cell_float(row[hmap["available"]])
                n += 1
            counts["investors"] = n

    # --- Collection ---
    if "Collection" in wb.sheetnames:
        ws = wb["Collection"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                cr = CollectionRow()
                if "source" in hmap:
                    x = row[hmap["source"]]
                    cr.source = str(x) if x is not None else None
                if "datewon" in hmap:
                    cr.date_won = _cell_date(row[hmap["datewon"]])
                plate = None
                if "platenumber" in hmap:
                    plate = row[hmap["platenumber"]]
                    cr.plate = str(plate).strip() if plate is not None else None
                    cr.plate_norm = norm_plate(cr.plate) or None
                if "make&model" in hmap:
                    x = row[hmap["make&model"]]
                    cr.model = str(x) if x is not None else None
                if "location" in hmap:
                    x = row[hmap["location"]]
                    cr.location = str(x) if x is not None else None
                if "postcode" in hmap:
                    x = row[hmap["postcode"]]
                    cr.post_code = str(x) if x is not None else None
                if "howfar?" in hmap or "howfar" in hmap:
                    hi = hmap.get("howfar?") or hmap.get("howfar")
                    if hi is not None:
                        x = row[hi]
                        cr.how_far = str(x) if x is not None else None
                if "collectiondate" in hmap:
                    x = row[hmap["collectiondate"]]
                    cr.collection_date = str(x) if x is not None else None
                if "number" in hmap:
                    x = row[hmap["number"]]
                    cr.number = str(x) if x is not None else None
                if "additionalnotes" in hmap:
                    x = row[hmap["additionalnotes"]]
                    cr.additional_notes = str(x) if x is not None else None
                if cr.plate_norm:
                    v = db.execute(select(Vehicle).where(Vehicle.plate_norm == cr.plate_norm)).scalar_one_or_none()
                    if v:
                        cr.stock_id = v.stock_id
                db.add(cr)
                n += 1
            counts["collections"] = n

    # --- Expense ---
    if "Expense" in wb.sheetnames:
        ws = wb["Expense"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                er = ExpenseRow()
                if "month" in hmap:
                    er.month = _month_date(row[hmap["month"]])
                if "date" in hmap:
                    er.date = _cell_date(row[hmap["date"]])
                if "category" in hmap:
                    x = row[hmap["category"]]
                    er.category = str(x) if x is not None else None
                fk = "from" if "from" in hmap else None
                if not fk:
                    fk = next((k for k in hmap if "from" in k or k.startswith("from")), None)
                if fk:
                    x = row[hmap[fk]]
                    er.from_vendor = str(x) if x is not None else None
                amt_key = "amount" if "amount" in hmap else next((k for k in hmap if "amount" in k), None)
                if amt_key:
                    er.amount = _cell_float(row[hmap[amt_key]])
                if "paymentmethod" in hmap:
                    x = row[hmap["paymentmethod"]]
                    er.payment_method = str(x) if x is not None else None
                if "paidby" in hmap:
                    x = row[hmap["paidby"]]
                    er.paid_by = str(x) if x is not None else None
                if "notes" in hmap:
                    x = row[hmap["notes"]]
                    er.notes = str(x) if x is not None else None
                db.add(er)
                n += 1
            counts["expenses"] = n

    # --- Money in / Money Out ---
    for sheet, model, extra in (
        ("Money in", MoneyInRow, True),
        ("Money Out", MoneyOutRow, False),
    ):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                if extra:
                    mr = MoneyInRow()
                    if "month" in hmap:
                        mr.month = _month_date(row[hmap["month"]])
                    if "date" in hmap:
                        mr.date = _cell_date(row[hmap["date"]])
                    if "category" in hmap:
                        x = row[hmap["category"]]
                        mr.category = str(x) if x is not None else None
                    ak = next((k for k in hmap if "amount" in k), None)
                    if ak:
                        mr.amount = _cell_float(row[hmap[ak]])
                    if "reg" in hmap:
                        x = row[hmap["reg"]]
                        mr.reg = str(x) if x is not None else None
                        mr.plate_norm = norm_plate(mr.reg) or None
                        if mr.plate_norm:
                            v = db.execute(select(Vehicle).where(Vehicle.plate_norm == mr.plate_norm)).scalar_one_or_none()
                            if v:
                                mr.stock_id = v.stock_id
                    if "notes" in hmap:
                        x = row[hmap["notes"]]
                        mr.notes = str(x) if x is not None else None
                    db.add(mr)
                else:
                    mr = MoneyOutRow()
                    if "month" in hmap:
                        mr.month = _month_date(row[hmap["month"]])
                    if "date" in hmap:
                        mr.date = _cell_date(row[hmap["date"]])
                    if "category" in hmap:
                        x = row[hmap["category"]]
                        mr.category = str(x) if x is not None else None
                    ak = next((k for k in hmap if "amount" in k), None)
                    if ak:
                        mr.amount = _cell_float(row[hmap[ak]])
                    if "notes" in hmap:
                        x = row[hmap["notes"]]
                        mr.notes = str(x) if x is not None else None
                    db.add(mr)
                n += 1
            counts[sheet.lower().replace(" ", "_")] = n

    # --- SOR ---
    if "SOR" in wb.sheetnames:
        ws = wb["SOR"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                sr = SorRow()
                if "month" in hmap:
                    sr.month = _month_date(row[hmap["month"]])
                if "dateaquired" in hmap or "dateacquired" in hmap:
                    ci = hmap.get("dateaquired") or hmap.get("dateacquired")
                    if ci is not None:
                        sr.date_acquired = _cell_date(row[ci])
                pk = "numberplatereference" if "numberplatereference" in hmap else "platenumber"
                if pk in hmap:
                    plate = row[hmap[pk]]
                    sr.plate = str(plate).strip() if plate is not None else None
                    sr.plate_norm = norm_plate(sr.plate) or None
                if "make&model" in hmap:
                    x = row[hmap["make&model"]]
                    sr.model = str(x) if x is not None else None
                if "sellername" in hmap:
                    x = row[hmap["sellername"]]
                    sr.seller_name = str(x) if x is not None else None
                if "totalcost" in hmap:
                    sr.total_cost = _cell_float(row[hmap["totalcost"]])
                if "saleprice" in hmap:
                    sr.sale_price = _cell_float(row[hmap["saleprice"]])
                if "breakdown" in hmap:
                    x = row[hmap["breakdown"]]
                    sr.breakdown = str(x) if x is not None else None
                if sr.plate_norm:
                    v = db.execute(select(Vehicle).where(Vehicle.plate_norm == sr.plate_norm)).scalar_one_or_none()
                    if v:
                        sr.stock_id = v.stock_id
                db.add(sr)
                n += 1
            counts["sor"] = n

    # --- Front Sheet ---
    if "Front Sheet" in wb.sheetnames:
        ws = wb["Front Sheet"]
        hmap, hri = _header_map(ws)
        if hmap and hri >= 0:
            n = 0
            for row in ws.iter_rows(min_row=hri + 2, values_only=True):
                if not row or all(x is None for x in row):
                    continue
                mk = "month" if "month" in hmap else None
                if not mk:
                    continue
                m = _month_date(row[hmap["month"]])
                if not m:
                    continue
                ms = db.execute(select(MonthlySummary).where(MonthlySummary.month == m)).scalar_one_or_none()
                if not ms:
                    ms = MonthlySummary(month=m)
                    db.add(ms)
                if "carssold" in hmap:
                    cs = row[hmap["carssold"]]
                    ms.cars_sold = int(cs) if cs is not None and str(cs).strip() != "" else None
                if "totalrevenue" in hmap:
                    ms.total_revenue = _cell_float(row[hmap["totalrevenue"]])
                if "totalgrossprofit" in hmap:
                    ms.total_gross_profit = _cell_float(row[hmap["totalgrossprofit"]])
                if "companyexpenses" in hmap:
                    ms.company_expenses = _cell_float(row[hmap["companyexpenses"]])
                if "totalsagrossprofit" in hmap:
                    ms.total_sa_gross_profit = _cell_float(row[hmap["totalsagrossprofit"]])
                if "investornetprofit" in hmap:
                    x = row[hmap["investornetprofit"]]
                    ms.investor_net_profit = str(x) if x is not None else None
                if "investorexpense" in hmap:
                    ms.investor_expense = _cell_float(row[hmap["investorexpense"]])
                if "companyfuelcosts" in hmap:
                    ms.company_fuel_costs = _cell_float(row[hmap["companyfuelcosts"]])
                if "othermoneyin" in hmap:
                    ms.other_money_in = _cell_float(row[hmap["othermoneyin"]])
                if "othermoneyout" in hmap:
                    ms.other_money_out = _cell_float(row[hmap["othermoneyout"]])
                if "netprofitexcinvestor" in hmap:
                    ms.net_profit_exc_investor = _cell_float(row[hmap["netprofitexcinvestor"]])
                if "netexcinvestor" in hmap:
                    ms.net_exc_investor = _cell_float(row[hmap["netexcinvestor"]])
                if "notes" in hmap:
                    x = row[hmap["notes"]]
                    ms.notes = str(x) if x is not None else None
                n += 1
            counts["front_sheet"] = n

    from dealershipos.services.excel_line_items import import_vehicle_line_sheets

    counts["vehicle_line_items"] = import_vehicle_line_sheets(wb, db)

    db.commit()
    wb.close()
    return counts


def _find_header_row(ws: Any, *needles: str) -> int:
    """First row whose cell contains any substring (case-insensitive)."""
    nl = tuple(n.lower() for n in needles)
    for r in range(1, 28):
        for c in range(1, 28):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).lower()
            if any(n in s for n in nl):
                return r
    return 1


def _write_sheet_replace_data_block(
    ws: Any,
    header_row: int,
    headers: list[str | None],
    rows: list[list[Any]],
) -> None:
    """Overwrite header + data in a fixed-width block; skip clearing formula cells."""
    ncols = len(headers)
    end_r = max(ws.max_row, header_row + len(rows) + 5)
    for r in range(header_row + 1, end_r + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            if getattr(cell, "data_type", None) == "f":
                continue
            cell.value = None
    for j, h in enumerate(headers, 1):
        if h is not None:
            ws.cell(header_row, j, h)
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            if j <= ncols:
                ws.cell(header_row + 1 + i, j, val)


def export_workbook(template_path: Path | str, db: Session, dest_path: Path | str) -> Path:
    """Copy a workbook and overwrite core sheets from the database (formula cells outside the block are left intact)."""
    template_path = Path(template_path)
    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    import shutil

    shutil.copy2(template_path, dest_path)
    wb = openpyxl.load_workbook(dest_path, data_only=False)

    # --- Vehicles: stock-like rows ---
    stock_headers = [
        "Stock ID",
        "Month",
        "Date Aquired",
        "Plate Number",
        "Make & Model",
        "Investor/SA",
        "Source",
        "PX Value",
        "Price",
        "Reconditioning costs",
        "Total Cost",
        "Sold",
        "Profit",
        "Status",
    ]
    stock_rows: list[list[Any]] = []
    for v in db.scalars(select(Vehicle).where(Vehicle.is_sold.is_(False))):
        stock_rows.append(
            [
                v.stock_id,
                v.month,
                v.date_acquired,
                v.plate,
                v.model,
                v.investor,
                v.source,
                v.px_value,
                v.purchase_price,
                v.recon_cost,
                v.total_cost,
                v.sold_price if v.sold_price else "-",
                v.profit if v.profit else "-",
                v.status,
            ]
        )
    if "Stock Data" in wb.sheetnames:
        ws = wb["Stock Data"]
        hr = _find_header_row(ws, "plate number", "stock id")
        _write_sheet_replace_data_block(ws, hr, stock_headers, stock_rows)

    sold_headers = [
        "Stock ID",
        "Month",
        "Date Aquired",
        "Number Plate reference",
        "Make & Model",
        "SA/Investor Name",
        "Total Cost",
        "Sold",
        "Part Ex",
        "SA/Investor Profit Share",
        "Total Profit",
        "Investor Profit",
        "SA Profit",
        "Date Listed",
        "Date Sold",
        "Days to Sell",
        "Platfrom",
        None,
        "Invoice Number",
        "Customer Name",
        "Contact info",
        "Warranty",
        "AutoGuard Number",
    ]
    sold_rows: list[list[Any]] = []
    for v in db.scalars(select(Vehicle).where(Vehicle.is_sold.is_(True))):
        sold_rows.append(
            [
                v.stock_id,
                v.month,
                v.date_acquired,
                v.plate,
                v.model,
                v.investor,
                v.total_cost,
                v.sold_price,
                v.part_ex,
                v.sa_investor_profit_share,
                v.profit,
                v.investor_profit,
                v.sa_profit,
                v.date_listed,
                v.date_sold,
                v.days_in_stock,
                v.platform,
                None,
                v.invoice_number,
                v.customer_name,
                v.contact_info,
                v.warranty,
                v.autoguard,
            ]
        )
    if "Sold Stock" in wb.sheetnames:
        ws = wb["Sold Stock"]
        hr = _find_header_row(ws, "number plate", "plate reference", "stock id")
        _write_sheet_replace_data_block(ws, hr, sold_headers, sold_rows)

    inv_headers = [
        "Investors",
        "Initial Balance",
        "Capital Returned",
        "Total Balance",
        "Purchased",
        "Total Profit (since Nov-25)",
        "Available",
    ]
    inv_rows = [
        [
            i.name,
            i.initial_balance,
            i.capital_returned,
            i.total_balance,
            i.purchased,
            i.total_profit,
            i.available,
        ]
        for i in db.scalars(select(Investor))
    ]
    if "Investor Budget" in wb.sheetnames:
        ws = wb["Investor Budget"]
        hr = _find_header_row(ws, "investors")
        _write_sheet_replace_data_block(ws, hr, inv_headers, inv_rows)

    col_headers = [
        "Source",
        "Date Won",
        "Plate Number",
        "Make & Model",
        "Location",
        "Post Code",
        "How Far?",
        "Collection Date",
        "Number",
        "Additional notes",
    ]
    col_rows = [
        [
            c.source,
            c.date_won,
            c.plate,
            c.model,
            c.location,
            c.post_code,
            c.how_far,
            c.collection_date,
            c.number,
            c.additional_notes,
        ]
        for c in db.scalars(select(CollectionRow))
    ]
    if "Collection" in wb.sheetnames:
        ws = wb["Collection"]
        hr = _find_header_row(ws, "plate number", "source")
        _write_sheet_replace_data_block(ws, hr, col_headers, col_rows)

    exp_headers = ["Month", "Date", "Category", "From", "Amount ", "Payment Method", "Paid By", "Notes"]
    exp_rows = [
        [
            e.month,
            e.date,
            e.category,
            e.from_vendor,
            e.amount,
            e.payment_method,
            e.paid_by,
            e.notes,
        ]
        for e in db.scalars(select(ExpenseRow))
    ]
    if "Expense" in wb.sheetnames:
        ws = wb["Expense"]
        hr = _find_header_row(ws, "category", "month")
        _write_sheet_replace_data_block(ws, hr, exp_headers, exp_rows)

    from dealershipos.services.excel_line_items import export_vehicle_line_sheets

    export_vehicle_line_sheets(wb, db)

    wb.save(dest_path)
    wb.close()
    return dest_path
